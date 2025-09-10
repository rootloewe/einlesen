import os
import glob
import csv
import base64
import pandas as pd
import xml.etree.ElementTree as ET
from pathlib import Path
from pikepdf import Pdf
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Alignment, Font
from decimal import Decimal
import tempfile


def pfad():
    #dateipfad = tempfile.gettempdir()
    dateipfad = r"/home/armin/projekte/github/einlesen"

    return Path(dateipfad)


def xltx_bauen(dateipfad, b64str):
    #xltx datei erstellen aus binärstr b64str
    xltx_datei = str(Path(dateipfad) / "vorlage.xltx")

    xls = open(xltx_datei, 'wb')
    xls.write(base64.b64decode(b64str))
    xls.close()


def xml_erstellen(dateipfad, pdf_files):
    # xml exportieren
    output_dir = dateipfad / "Daten" / "xml"
    os.makedirs(output_dir, exist_ok=True)

    xml_counter = 1

    for pdf in pdf_files:
        try:
            pdf_obj = Pdf.open(pdf)
            if hasattr(pdf_obj, "attachments"):
                attachments = pdf_obj.attachments
                if attachments:
                    print(f"\nAnhänge in {pdf}:")

                    for name, obj in attachments.items():
                        print(f"  - {name}")
                        # Prüfe, ob der Anhang eine XML ist
                        if name.lower().endswith(".xml"):
                            out_file = os.path.join(output_dir, f"{xml_counter}.xml")
                        else:
                            print(f"Kein xml im Anhang der pdf {pdf}")
                            continue
                        try:
                            if (hasattr(obj, "obj") and "/EF" in obj.obj and "/F" in obj.obj["/EF"]):
                                embedded_file = obj.obj["/EF"]["/F"]

                                if hasattr(embedded_file, "get_object"):
                                    embedded_obj = embedded_file.get_object()
                                else:
                                    embedded_obj = embedded_file

                                if hasattr(embedded_obj, "get_data"):
                                    data = embedded_obj.get_data()
                                elif hasattr(embedded_obj, "read_bytes"):
                                    data = embedded_obj.read_bytes()
                                else:
                                    data = bytes(embedded_obj)

                                datei = open(out_file, "wb")
                                datei.write(data)

                                print(f"    -> {out_file}")
                                print(f"Datei erstellt: {os.path.abspath(out_file)}")
                                xml_counter += 1
                                pdf_obj.close()
                                datei.close()
                            else:
                                print(f"Fehler: Kein Embedded-File-Stream gefunden für {name}")
                        except Exception as e:
                            print(f"Fehler beim Extrahieren von {name}: {e}")
                    print("Extrahiert.")
                else:
                    print(f"{pdf}: Keine Anhänge gefunden.")
            else:
                print(f"{pdf}: Keine Anhänge gefunden (kein attachments-Attribut).")
        except Exception as e:
            print(f"Fehler bei {pdf}: {e}")

    # weil wenn beim zweiten Durchlauf, weniger Rechnungen existieren könnten, bleiben alte Daten, die dann auch mit eingelesen werden
    # daher werden diese gelöscht
    while True:
        file_path = os.path.join(output_dir, f"{xml_counter}.xml")

        if os.path.exists(file_path):
            try:
                os.remove(file_path)
                print(f"Datei {file_path} wurde erfolgreich gelöscht.")
            except FileNotFoundError:
                print(f"Datei {file_path} wurde nicht gefunden.")
            except PermissionError:
                print(f"Keine Berechtigung, um die Datei {file_path} zu löschen.")
            except Exception as e:
                print(f"Fehler beim Löschen der Datei: {e}")
        else:
            break

        xml_counter += 1


def csv_erstellen(dateipfad):
    # csv ERstellen
    # Verzeichnis mit den XML-Dateien
    xml_verzeichnis = str(dateipfad / "Daten" / "xml")
    csv_datei = str(dateipfad / "Daten" / "Rechnungen.csv")

    # XML-Namespace-Definitionen
    ns = {
        "rsm": "urn:ferd:CrossIndustryDocument:invoice:1p0",
        "ram": "urn:un:unece:uncefact:data:standard:ReusableAggregateBusinessInformationEntity:12",
        "udt": "urn:un:unece:uncefact:data:standard:UnqualifiedDataType:15",
    }

    # Alle XML-Dateien im Verzeichnis finden
    xml_dateien = glob.glob(os.path.join(xml_verzeichnis, "*.xml"))

    if not xml_dateien:
        print("Keine xml Dateien gefunden.")
        return False


    csv_obj = open(csv_datei, mode="w", newline="", encoding="utf-8")
    writer = csv.writer(csv_obj, delimiter=";")

    for xml_datei in xml_dateien:
        try:
            tree = ET.parse(xml_datei)
            root = tree.getroot()

            # Rechnungsnummer
            rechnungsnummer_elem = root.find(".//rsm:HeaderExchangedDocument/ram:ID", ns)
            rechnungsnummer = (rechnungsnummer_elem.text if rechnungsnummer_elem is not None else "")

            # Betrag
            betrag_elem = root.find(".//ram:SpecifiedTradeSettlementMonetarySummation/ram:GrandTotalAmount", ns,)
            betrag = betrag_elem.text if betrag_elem is not None else "0.00"

            # Skonto-Datum
            skonto_datum_elem = root.find(".//ram:ApplicableTradePaymentDiscountTerms/ram:BasisDateTime/udt:DateTimeString", ns,)

            if skonto_datum_elem is not None and skonto_datum_elem.text:
                skonto_datum_raw = skonto_datum_elem.text
                skonto_datum = f"{skonto_datum_raw[:4]}-{skonto_datum_raw[4:6]}-{skonto_datum_raw[6:]}"
            else:
                skonto_datum = ""

            # Skontobetrag
            skontobetrag_elem = root.find(".//ram:ApplicableTradePaymentDiscountTerms/ram:ActualDiscountAmount", ns,)
            skontobetrag = (skontobetrag_elem.text if skontobetrag_elem is not None else "0.00")

            # Überweisen berechnen (Betrag - Skontobetrag)
            try:
                ueberweisen = Decimal(betrag.replace(",", ".")) - Decimal(skontobetrag.replace(",", "."))
                ueberweisen = f"{ueberweisen:.2f}"
            except Exception:
                ueberweisen = ""

            betrag_f = Decimal(betrag.replace(",", "."))
            skontobetrag_f = Decimal(skontobetrag.replace(",", "."))
            ueberweisen_f = Decimal(ueberweisen)

            # In ein datetime-Objekt umwandeln
            datum_dt = datetime.strptime(skonto_datum, "%Y-%m-%d")
            datum_formatiert = datum_dt.strftime("%d.%m.%Y")

            # Zeile schreiben
            writer.writerow([datum_formatiert, rechnungsnummer, betrag_f, skontobetrag_f, "", ueberweisen_f, ])

        except Exception as e:
            print(f"Fehler bei {xml_datei}: {e}")

    csv_obj.close()

    # umdie Tablle vor Excelimport nach Datum zu sortieren
    def datum_sortieren(zeile):
        try:
            return datetime.strptime(zeile[0], "%d.%m.%Y")
        except ValueError:
            return datetime.min

    csv_obj = open(csv_datei, mode="r", newline="", encoding="utf-8")
    reader = csv.reader(csv_obj, delimiter=";")
    # header = next(reader)
    daten = list(reader)

    daten.sort(key=datum_sortieren)
    csv_obj.close()

    # Zurückschreiben
    csv_obj = open(csv_datei, mode="w", newline="", encoding="utf-8")
    writer = csv.writer(csv_obj, delimiter=";")
    # writer.writerow(header)

    for i, zeile in enumerate(daten, start=1):
        writer.writerow([i] + zeile)

    csv_obj.close()

    print(f"Alle Daten wurden in {csv_datei} gespeichert.")

    return True


def xltx_erstellen(dateipfad):
    # In xltx templade schreiben
    # CSV einlesen
    csv_datei = str(dateipfad / "Daten" / "Rechnungen.csv")
    xltx_datei = str(dateipfad / "vorlage.xltx")

    df = pd.read_csv(csv_datei, sep=';', decimal='.', header=None)

    try:
        wb = load_workbook(xltx_datei)
        ws = wb.worksheets[0]
    except FileNotFoundError:
        print(f"Die Datei {xltx_datei} existiert nicht.")
    except Exception as e:
        print(f"Fehler beim Laden der Datei: {e}")


    # Löschen der alten Werte. unnötig, da nun mit binärstr gearbeitet wird
    # for row in range(13, 150):  # Zeilen 13 bis 150 inkl.
    #     for col in range(1, 9):  # Spalte 1 (B) bis 8 (H)
    #         cell = ws.cell(row=row, column=col)
    #         cell.value = None
    #         cell.fill = PatternFill(fill_type=None)
    #         cell.alignment = Alignment(horizontal='right')
    #         if col == 2 or col == 3:
    #             cell.alignment = Alignment(horizontal='center')


    # DataFrame ab Zelle A13 schreiben
    start_row = 13
    start_col = 1

    for i, row in enumerate(df.values):
        for j, value in enumerate(row):
            # Versuche, Wert als Float zu speichern, sonst als String
            try:
                ws.cell(row=start_row + i, column=start_col + j, value=Decimal(str(value).replace(',', '.')))
            except Exception:
                ws.cell(row=start_row + i, column=start_col + j, value=str(value))


    #Fußzeile schreiben
    letzte_zeile = len(df.values) - 1 + start_row

    grün = PatternFill(start_color='a9d18e', end_color='a9d18e', fill_type='solid')
    rot = PatternFill(start_color='ff0000', end_color='ff0000', fill_type='solid')

    for col in range(2, 9):
        ws.cell(row=letzte_zeile + 2, column=col).fill = grün

    for col in range(4, 8):
        ws.cell(row=letzte_zeile + 4, column=col).fill = rot

    zelle = [None] * 3
    zelle[0] = ws.cell(row=letzte_zeile + 2, column=4, value="ges. Betrag",)
    zelle[1] = ws.cell(row=letzte_zeile + 2, column=5, value="ges. Skonto",)
    zelle[2] = ws.cell(row=letzte_zeile + 2, column=7, value="Gesamt",)

    for i, wert in enumerate(zelle):
        wert.alignment = Alignment(horizontal='center', vertical='center')


    df[df.columns[3]] = df[df.columns[3]].apply(lambda x: Decimal(str(x)))
    erste_sum = df[df.columns[3]].sum()
    df[df.columns[4]] = df[df.columns[4]].apply(lambda x: Decimal(str(x)))
    zweite_sum = df[df.columns[4]].sum()
    df[df.columns[6]] = df[df.columns[6]].apply(lambda x: Decimal(str(x)))
    dritte_sum = df[df.columns[6]].sum()

    #Hier schreibe ich die Formeln in Excel und dort wird gerechnet
    erste_formel = f"=SUM(D13:D{letzte_zeile+1})"
    zweite_formel = f"=SUM(E13:E{letzte_zeile+1})"
    dritte_formel = f"=SUM(G13:G{letzte_zeile+1})"


    summen = [None] * 3
    summen[0] = ws.cell(row=letzte_zeile + 4, column=4, value=erste_formel)
    summen[1] = ws.cell(row=letzte_zeile + 4, column=5, value=zweite_formel)
    summen[2] = ws.cell(row=letzte_zeile + 4, column=7, value=dritte_formel)

    ws.cell(row=4, column=4, value=f"=G{letzte_zeile+4}")

    for i, wert in enumerate(summen):
        wert.alignment = Alignment(horizontal="right",)


    ws = wb.worksheets[0]
    ws.cell(row=160, column=12, value=erste_sum).font = Font(color='FFFFFF')
    ws.cell(row=160, column=13, value=zweite_sum).font = Font(color='FFFFFF')
    ws.cell(row=160, column=14, value=dritte_sum).font = Font(color='FFFFFF')

    wb.save(xltx_datei)

    print("CSV-Daten wurden in die ODS-Datei geschrieben.")


def aufräumen(dateiname):
    try:
        Path(dateiname).unlink()
    except FileNotFoundError:
        print("Datei nicht gefunden.", dateiname)
    except PermissionError:
        print("Keine Berechtigung zum Löschen der Datei.")


